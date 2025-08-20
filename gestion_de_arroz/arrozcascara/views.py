from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
# Create your views here.
from django.contrib import messages
from .models import Representante, Factura
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import json
from django.db.models import Sum
import pandas as pd
from django.utils.timezone import now
from django.views.decorators.http import require_http_methods
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from django.utils.dateparse import parse_date
from datetime import datetime, date
from django.db import transaction


def representantes(request):
    return render(request, "arrozcascara/representantes.html")

def registrar_representante(request):
    if request.method == 'POST':
        cedula = request.POST.get('cedula')
        nombre_completo = request.POST.get('nombre_completo')
        direccion = request.POST.get('direccion')
        
        try:
            # Crear y guardar el nuevo representante
            Representante.objects.create(
                cedula=cedula,
                nombre_completo=nombre_completo,
                direccion=direccion
            )
            messages.success(request, 'Representante registrado exitosamente')
        except Exception as e:
            messages.error(request, f'Error al registrar representante: {str(e)}')
        
        return redirect('representantes')

def index(request):
    return render(request, "arrozcascara/index.html")

def dashboard(request):
    # Obtener datos de la base de datos
    representantes = Representante.objects.all().values('id', 'nombre_completo', 'cedula', 'direccion')
    facturas = Factura.objects.all().values('representante', 'cantidad_sacos', 'fecha')
    
    # Convertir a DataFrames de pandas
    df_representantes = pd.DataFrame(list(representantes))
    df_facturas = pd.DataFrame(list(facturas))
    
    # Si no hay datos, usar valores por defecto
    if df_representantes.empty or df_facturas.empty:
        context = {
            'total_representantes': 0,
            'total_facturas': 0,
            'total_sacos': 0,
            'promedio_sacos': 0,
            'representantes_data': [],
            'chart_labels': [],
            'chart_sacks_data': [],
            'chart_invoices_data': [],
            'chart_colors': [],
            'last_update': '--:--'
        }
        return render(request, "arrozcascara/dashboard.html", context)
    
    # Procesar datos con pandas
    df_facturas['cantidad_sacos'] = pd.to_numeric(df_facturas['cantidad_sacos'])
    
    # Agrupar datos por representante
    df_agrupado = df_facturas.groupby('representante').agg(
        total_sacos=('cantidad_sacos', 'sum'),
        total_facturas=('cantidad_sacos', 'count')
    ).reset_index()
    
    # Combinar con datos de representantes
    df_final = pd.merge(
        df_representantes, 
        df_agrupado, 
        left_on='id', 
        right_on='representante'
    )
    
    # Ordenar por total de sacos (descendente)
    df_final = df_final.sort_values('total_sacos', ascending=False)
    
    # Calcular estadísticas generales
    total_representantes = df_final.shape[0]
    total_facturas = df_final['total_facturas'].sum()
    total_sacos = df_final['total_sacos'].sum()
    promedio_sacos = round(total_sacos / total_representantes) if total_representantes > 0 else 0
    
    # Preparar datos para gráficos
    chart_labels = df_final['nombre_completo'].tolist()
    chart_sacks_data = df_final['total_sacos'].tolist()
    chart_invoices_data = df_final['total_facturas'].tolist()
    
    # Colores para los gráficos
    chart_colors = [
        '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
        '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9'
    ]
    
    # Obtener hora de actualización
    last_update = now().strftime('%H:%M')
    
    context = {
        'total_representantes': total_representantes,
        'total_facturas': total_facturas,
        'total_sacos': total_sacos,
        'promedio_sacos': promedio_sacos,
        'representantes_data': df_final.to_dict('records'),
        'chart_labels': chart_labels,
        'chart_sacks_data': chart_sacks_data,
        'chart_invoices_data': chart_invoices_data,
        'chart_colors': chart_colors,
        'last_update': last_update
    }
    
    return render(request, "arrozcascara/dashboard.html", context)



def detalles(request):
    try:
        # Obtener parámetros de filtrado con valores por defecto
        representante_id = request.GET.get('representante', '')
        variedad = request.GET.get('variedad', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        
        # Debug: Mostrar parámetros recibidos
        print("\n=== Parámetros de filtrado ===")
        print(f"Representante: {representante_id}")
        print(f"Variedad: {variedad}")
        print(f"Fecha desde: {fecha_desde}")
        print(f"Fecha hasta: {fecha_hasta}")
        
        # Validar fechas si ambas están presentes
        date_error = None
        if fecha_desde and fecha_hasta:
            if fecha_desde > fecha_hasta:
                date_error = 'La fecha "Desde" no puede ser mayor que la fecha "Hasta"'
                messages.error(request, date_error)
                # Intercambiar fechas para corregir automáticamente
                fecha_desde, fecha_hasta = fecha_hasta, fecha_desde
        
        # Inicializar queryset con select_related para optimización
        facturas = Factura.objects.all().select_related('representante').order_by('representante__nombre_completo', '-fecha')
        
        # Aplicar filtros si existen
        if representante_id:
            facturas = facturas.filter(representante_id=representante_id)
        
        if variedad:
            facturas = facturas.filter(variedad__iexact=variedad)  # Búsqueda case-insensitive
        
        if fecha_desde:
            facturas = facturas.filter(fecha__gte=fecha_desde)
        
        if fecha_hasta:
            facturas = facturas.filter(fecha__lte=fecha_hasta)
        
        # Obtener todos los representantes (para el dropdown de filtros)
        representantes = Representante.objects.all().order_by('nombre_completo')
        
        # Agrupar facturas por representante con total de sacos
        facturas_por_representante = []
        representantes_con_facturas = set(facturas.values_list('representante_id', flat=True))
        
        for rep in representantes:
            if rep.id in representantes_con_facturas:
                facturas_rep = facturas.filter(representante=rep)
                total_sacos_rep = facturas_rep.aggregate(total=Sum('cantidad_sacos'))['total'] or 0
                
                facturas_por_representante.append({
                    'representante': rep,
                    'facturas': facturas_rep,
                    'total_sacos': total_sacos_rep
                })
        
        # Calcular totales generales
        total_representantes = len(facturas_por_representante)
        total_facturas = facturas.count()
        total_sacos = facturas.aggregate(total=Sum('cantidad_sacos'))['total'] or 0
        
        # Debug: Mostrar resultados del filtrado
        print("\n=== Resultados del filtrado ===")
        print(f"Total representantes con facturas: {total_representantes}")
        print(f"Total facturas: {total_facturas}")
        print(f"Total sacos: {total_sacos}")
        
        for grupo in facturas_por_representante:
            print(f"\nRepresentante: {grupo['representante'].nombre_completo}")
            print(f"Total sacos: {grupo['total_sacos']}")
            print(f"Facturas: {grupo['facturas'].count()}")
            for factura in grupo['facturas']:
                print(f"  - Factura: {factura.numero_factura} | Cliente: {factura.nombre_cliente} | Sacos: {factura.cantidad_sacos} | Fecha: {factura.fecha}")
        
        # Preparar contexto
        context = {
            'representantes': representantes,
            'facturas': facturas,
            'facturas_por_representante': facturas_por_representante,
            'total_representantes': total_representantes,
            'total_facturas': total_facturas,
            'total_sacos': total_sacos,
            'request': request,  # Para acceder a los parámetros GET en el template
            'filtros_aplicados': {
                'representante': representante_id,
                'variedad': variedad,
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta
            },
            'date_error': date_error
        }
        
        return render(request, "arrozcascara/detalles.html", context)
    
    except Exception as e:
        error_msg = f"Error al cargar los datos: {str(e)}"
        print(f"\n=== ERROR en vista detalles ===\n{error_msg}\n")
        messages.error(request, error_msg)
        
        # Retornar contexto básico en caso de error
        return render(request, "arrozcascara/detalles.html", {
            'representantes': Representante.objects.all().order_by('nombre_completo'),
            'facturas_por_representante': [],
            'total_representantes': 0,
            'total_facturas': 0,
            'total_sacos': 0,
            'request': request,
            'filtros_aplicados': {},
            'error': error_msg
        })



@require_http_methods(["GET"])
def obtener_factura(request, invoice_id):
    try:
        factura = Factura.objects.get(id=invoice_id)
        return JsonResponse({
            'success': True,
            'factura': {
                'id': factura.id,
                'numero_factura': factura.numero_factura,
                'cedula': factura.cedula,
                'nombre_cliente': factura.nombre_cliente,
                'cantidad_sacos': factura.cantidad_sacos,
                'representante_id': factura.representante_id,
                'fecha': factura.fecha.strftime('%Y-%m-%d'),
                'variedad': factura.variedad
            }
        })
    except Factura.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Factura no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_http_methods(["POST"])
def editar_factura(request, invoice_id):
    try:
        factura = Factura.objects.get(id=invoice_id)
        
        # Leer datos del FormData (no JSON)
        numero_factura = request.POST.get('numero_factura')
        cedula = request.POST.get('cedula')
        nombre_cliente = request.POST.get('nombre_cliente')
        cantidad_sacos = request.POST.get('cantidad_sacos')
        representante_id = request.POST.get('representante_id')
        fecha = request.POST.get('fecha')
        variedad = request.POST.get('variedad')
        
        # Validación y conversión de cantidad_sacos
        try:
            cantidad_sacos = int(cantidad_sacos) if cantidad_sacos else 0
            if cantidad_sacos <= 0:
                return JsonResponse({'success': False, 'error': 'La cantidad de sacos debe ser mayor a 0'}, status=400)
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'La cantidad de sacos debe ser un número válido'}, status=400)
        
        # Validación de representante
        if representante_id:
            try:
                representante = Representante.objects.get(id=representante_id)
            except Representante.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Representante no encontrado'}, status=400)
        
        # Actualización de campos
        if numero_factura:
            factura.numero_factura = numero_factura
        if cedula:
            factura.cedula = cedula
        if nombre_cliente:
            factura.nombre_cliente = nombre_cliente
        if cantidad_sacos:
            factura.cantidad_sacos = cantidad_sacos
        if representante_id:
            factura.representante_id = representante_id
        if fecha:
            factura.fecha = fecha
        if variedad:
            factura.variedad = variedad
        
        factura.save()
        
        return JsonResponse({'success': True, 'message': 'Factura actualizada correctamente'})
        
    except Factura.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Factura no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'}, status=500)



@require_http_methods(["POST"])
def eliminar_factura(request, invoice_id):
    try:
        factura = Factura.objects.get(id=invoice_id)
        factura.delete()
        return JsonResponse({'success': True})
    except Factura.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Factura no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@require_http_methods(["POST"])
def pagar_factura(request, invoice_id):
    try:
        factura = Factura.objects.get(id=invoice_id)
        
        # Validar y convertir monto
        try:
            monto = Decimal(request.POST.get('monto', '0'))
            if monto <= 0:
                return JsonResponse({'success': False, 'error': 'El monto debe ser mayor a 0'}, status=400)
        except (InvalidOperation, TypeError):
            return JsonResponse({'success': False, 'error': 'Monto no válido'}, status=400)
        
        # Validar fecha de pago
        fecha_pago = request.POST.get('fecha_pago')
        if not fecha_pago:
            return JsonResponse({'success': False, 'error': 'Fecha de pago requerida'}, status=400)
        
        # Actualizar factura
        factura.monto = monto
        factura.estado = 'pagado'
        factura.fecha_pago = fecha_pago
        factura.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Pago de ${monto} registrado para factura {factura.numero_factura}'
        })
        
    except Factura.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Factura no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)




def registrodefacturas(request):
    representantes = Representante.objects.all().order_by('nombre_completo')
    return render(request, "arrozcascara/registrodefacturas.html", {
        'representantes': representantes
    })



@require_http_methods(["GET", "POST"])
def registro_facturas(request):
    representantes = Representante.objects.all().order_by('nombre_completo')
    
    if request.method == 'GET':
        return render(request, "arrozcascara/registrodefacturas.html", {
            'representantes': representantes
        })
    
    # POST handling
    try:
        if 'excel-file' in request.FILES:
            return handle_excel_upload(request)
        return handle_manual_form(request)
    except Exception as e:
        messages.error(request, f'Error en el proceso: {str(e)}')
        return redirect('registro_facturas')




@require_POST
def registrar_factura(request):
    try:
        # Verificar si es carga de Excel
        if 'excel-file' in request.FILES or request.POST.get('form_type') == 'excel':
            return handle_excel_upload(request)
        
        # Manejar formulario manual
        return handle_manual_form(request)
    
    except Exception as e:
        messages.error(request, f'Error en el proceso: {str(e)}')
        return redirect('registrodefacturas')




def handle_excel_upload(request):
    try:
        excel_file = request.FILES['excel-file']
        
        # Usar with para asegurar que el archivo se cierre correctamente
        with transaction.atomic():  # Usar transacción atómica
            wb = load_workbook(excel_file, read_only=True)
            sheet = wb.active
            
            # Validar estructura del archivo
            if sheet.max_row < 2:
                messages.error(request, 'El archivo Excel no contiene datos')
                return redirect('registro_facturas')
            
            # Obtener encabezados
            headers = [cell.value for cell in sheet[1]]
            
            # Mapeo de columnas mejorado
            column_indices = {
                'numero_factura': next((i for i, h in enumerate(headers) 
                                      if h and any(x in str(h).lower() 
                                      for x in ['factura', 'numero', 'n°', 'no'])), None),
                'cedula': next((i for i, h in enumerate(headers) 
                              if h and any(x in str(h).lower() 
                              for x in ['cedula', 'cédula', 'id'])), None),
                # ... agregar mapeos para los demás campos
            }
            
            # Verificar campos obligatorios
            required = ['numero_factura', 'cedula', 'nombre', 'sacos', 'variedad']
            missing = [field for field in required if column_indices.get(field) is None]
            
            if missing:
                messages.error(request, f'Faltan columnas requeridas: {", ".join(missing)}')
                return redirect('registro_facturas')
            
            # Procesar filas
            success_count = 0
            errors = []
            
            for row_idx in range(2, sheet.max_row + 1):
                row = [cell.value for cell in sheet[row_idx]]
                try:
                    # Extraer datos
                    factura_data = {
                        'numero_factura': str(row[column_indices['numero_factura']]).strip(),
                        'cedula': str(row[column_indices['cedula']]).strip(),
                        'nombre_cliente': str(row[column_indices['nombre']]).strip(),
                        'cantidad_sacos': int(float(row[column_indices['sacos']])),
                        'variedad': str(row[column_indices['variedad']]).strip(),
                        'fecha': parse_excel_date(row[column_indices.get('fecha')]),
                        'estado': 'pendiente'
                    }
                    
                    # Obtener representante
                    rep_col = column_indices.get('representante')
                    if rep_col is not None and row[rep_col]:
                        try:
                            factura_data['representante'] = Representante.objects.get(
                                nombre_completo__icontains=str(row[rep_col]).strip()
                            )
                        except (Representante.DoesNotExist, Representante.MultipleObjectsReturned):
                            factura_data['representante'] = Representante.objects.first()
                    else:
                        factura_data['representante'] = Representante.objects.first()
                    
                    # Validar unicidad de número de factura
                    if Factura.objects.filter(numero_factura=factura_data['numero_factura']).exists():
                        raise ValueError(f"La factura {factura_data['numero_factura']} ya existe")
                    
                    # Crear y guardar la factura
                    Factura.objects.create(**factura_data)
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Fila {row_idx}: {str(e)}")
                    continue
            
            # Resultados
            if success_count > 0:
                messages.success(request, f'Se importaron {success_count} facturas correctamente')
            
            if errors:
                error_msg = f"Errores en {len(errors)} filas"
                if len(errors) <= 5:
                    error_msg += ": " + " | ".join(errors)
                else:
                    error_msg += f". Primeros errores: {' | '.join(errors[:5])}... (+{len(errors)-5} más)"
                messages.warning(request, error_msg)
            
            return redirect('registro_facturas')
    
    except Exception as e:
        messages.error(request, f'Error al procesar el archivo: {str(e)}')
        return redirect('registro_facturas')



def handle_manual_form(request):
    try:
        numero_factura = request.POST.get('numeroFactura')
        cedula = request.POST.get('cedula')
        nombre_cliente = request.POST.get('nombre')
        cantidad_sacos = request.POST.get('cantidad_sacos')
        representante_id = request.POST.get('representante_id')
        fecha = request.POST.get('fecha')
        variedad = request.POST.get('variedad')

        # Validaciones básicas
        if not all([numero_factura, cedula, nombre_cliente, cantidad_sacos, representante_id, fecha, variedad]):
            messages.error(request, 'Todos los campos son requeridos')
            return redirect('registrodefacturas')
        
        # Validar que la factura no exista
        if Factura.objects.filter(numero_factura=numero_factura).exists():
            messages.error(request, f'Ya existe una factura con el número {numero_factura}')
            return redirect('registrodefacturas')
        
        representante = Representante.objects.get(id=representante_id)

        Factura.objects.create(
            numero_factura=numero_factura,
            cedula=cedula,
            nombre_cliente=nombre_cliente,
            cantidad_sacos=int(cantidad_sacos),
            representante=representante,
            fecha=fecha,
            variedad=variedad,
            estado='pendiente'
        )
        
        messages.success(request, 'Factura registrada exitosamente')
        return redirect('registrodefacturas')
    
    except Representante.DoesNotExist:
        messages.error(request, 'El representante seleccionado no existe')
    except ValueError:
        messages.error(request, 'La cantidad de sacos debe ser un número válido')
    except Exception as e:
        messages.error(request, f'Error al registrar factura: {str(e)}')
    
    return redirect('registrodefacturas')

def get_cell_value(row, column_index):
    """Obtiene el valor de una celda de manera segura"""
    if column_index is None or column_index >= len(row):
        return None
    
    cell = row[column_index]
    value = cell.value
    
    if value is None:
        return None
    
    # Convertir a string y limpiar espacios
    return str(value).strip()

def parse_excel_date(excel_date):
    """Convierte fechas de Excel a formato Python"""
    if excel_date is None:
        return date.today()
    
    try:
        # Si es un número (fecha de Excel)
        if isinstance(excel_date, (int, float)):
            from datetime import timedelta
            return (datetime(1899, 12, 30) + timedelta(days=excel_date)).date()
        
        # Si es una fecha datetime
        elif hasattr(excel_date, 'date'):
            return excel_date.date()
        
        # Si es un string, intentar parsearlo
        elif isinstance(excel_date, str):
            excel_date = excel_date.strip()
            
            # Intentar diferentes formatos
            formats = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']
            for fmt in formats:
                try:
                    return datetime.strptime(excel_date, fmt).date()
                except ValueError:
                    continue
            
            # Último intento con el parser de Django
            parsed_date = parse_date(excel_date)
            if parsed_date:
                return parsed_date
        
        # Si no se puede parsear, usar fecha actual
        return date.today()
    
    except Exception:
        return date.today()



def gestionderepresentantes(request):
    representantes = Representante.objects.all().order_by('nombre_completo')
    return render(request, "arrozcascara/gestionderepresentantes.html", {
        'representantes': representantes
    })

def get_representante(request, id):
    try:
        representante = get_object_or_404(Representante, id=id)
        data = {
            'success': True,
            'representante': {
                'id': representante.id,
                'cedula': representante.cedula,
                'nombre_completo': representante.nombre_completo,
                'direccion': representante.direccion
            }
        }
    except Exception as e:
        data = {
            'success': False,
            'message': str(e)
        }
    return JsonResponse(data)

@require_POST
def editar_representante(request, id):
    try:
        representante = get_object_or_404(Representante, id=id)
        representante.cedula = request.POST.get('cedula')
        representante.nombre_completo = request.POST.get('nombre_completo')
        representante.direccion = request.POST.get('direccion')
        representante.save()
        
        messages.success(request, 'Representante actualizado exitosamente')
    except Exception as e:
        messages.error(request, f'Error al actualizar representante: {str(e)}')
    
    return redirect('gestionderepresentantes')

@require_POST
def eliminar_representante(request, id):
    try:
        representante = get_object_or_404(Representante, id=id)
        representante.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Representante eliminado exitosamente'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al eliminar representante: {str(e)}'
        }, status=500)










from datetime import datetime, date, timedelta
from django.utils.dateparse import parse_date

def parse_excel_date(excel_date):
    """Convierte varios formatos de fecha de Excel a objeto date"""
    if excel_date is None:
        return date.today()
    
    # Si es número de Excel (días desde 1900-01-01)
    if isinstance(excel_date, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=excel_date)).date()
        except:
            return date.today()
    
    # Si ya es un objeto datetime
    if isinstance(excel_date, datetime):
        return excel_date.date()
    
    # Si es string, intentar parsear
    if isinstance(excel_date, str):
        try:
            return parse_date(excel_date.strip())
        except (ValueError, TypeError):
            pass
    
    return date.today()